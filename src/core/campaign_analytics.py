"""
campaign_analytics.py — 스마트스토어 마케팅분석 표(붙여넣기)를 파싱·집계하는 순수 함수.

streamlit 의존 없음(테스트 가능). 입력은 스마트스토어 '마케팅분석'에서 표를 복사한
raw 텍스트(탭 구분, 헤더 여러 줄 포함 가능).

컬럼 고정 인덱스(검증됨):
  0=채널속성(전체/모바일/PC)  2=nt_medium  3=nt_detail  4=nt_keyword
  6=유입수  9=결제수(마지막클릭)  11=결제금액(마지막클릭)

데이터행 판별: 탭 분리 후 len>=12 이고 col[0] in {전체,모바일,PC} 이며 col[6]이 숫자인 줄만.
  (헤더가 몇 줄이든 이 규칙이 자동으로 걸러낸다.)

집계 원칙:
  · summary = '전체' 행 합(기기 전체 그랜드토탈). '전체' 행이 없으면 기기(모바일/PC)행 합.
  · by_medium / by_campaign / by_product = ★기기행(모바일+PC)만 합산('전체'는 그랜드토탈이라
    중복집계 방지로 제외). by_campaign 은 nt_detail 로 모바일+PC 를 병합한다.
"""

from __future__ import annotations

import re

# 성과 임계값(결제율 %) — 조정 가능.
PERF_GOOD_RATE = 15.0
PERF_BAD_RATE = 5.0

_DEVICE_ALL = "전체"
_DEVICE_SPLIT = ("모바일", "PC")
_CHANNELS = {_DEVICE_ALL, *_DEVICE_SPLIT}

# 컬럼 인덱스.
_C_CHANNEL, _C_MEDIUM, _C_DETAIL, _C_KEYWORD = 0, 2, 3, 4
_C_INFLOW, _C_PAY, _C_AMOUNT = 6, 9, 11
# +14일 기여도추정(마지막클릭 9/11 다음) — 보조 지표(결제수·결제금액만).
_C_PAY14, _C_AMOUNT14 = 13, 15
_MIN_COLS = 12

# 구형 detail → 제품 추정용 키워드(부분일치, 더 구체적인 것 먼저).
_PRODUCT_SUBSTR = [
    (("에어로닷", "pump", "펌프"), "에어로닷"),
    (("거치대", "holder"), "거치대"),
    (("네비", "navi", "필름", "film"), "네비필름"),
    (("유리복원", "glass"), "유리복원제"),
    (("와이퍼", "wiper"), "와이퍼"),
    (("필터", "filter"), "에어컨필터"),
]


def _num(s: str) -> float:
    """콤마·% 제거 후 float. 비거나 못 읽으면 0.0."""
    s = (s or "").replace(",", "").replace("%", "").strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _is_num(s: str) -> bool:
    """콤마·% 제거 후 숫자로 읽히면 True(빈 문자열은 False)."""
    s = (s or "").replace(",", "").replace("%", "").strip()
    if not s:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def _rate(pay: float, inflow: float) -> float:
    """결제율(%) — 유입 0 방어."""
    return (pay / inflow * 100.0) if inflow else 0.0


def _perf_tag(rate: float) -> str:
    if rate >= PERF_GOOD_RATE:
        return "good"
    if rate < PERF_BAD_RATE:
        return "bad"
    return "mid"


def _is_new_detail(detail: str) -> bool:
    """신규형 detail(`날짜_제품_차종`) — '_'로 3토큰 이상."""
    return detail.count("_") >= 2


def _detect_product(detail: str) -> str:
    """구형 detail 에서 제품 키워드 substring 감지. 못 찾으면 '미분류'."""
    low = detail.lower()
    for needles, product in _PRODUCT_SUBSTR:
        if any(n.lower() in low for n in needles):
            return product
    return "미분류"


_C_SOURCE = 1


class _Row:
    __slots__ = (
        "channel", "source", "medium", "detail", "keyword", "inflow", "pay", "amount",
        "pay14", "amount14")

    def __init__(self, cols: list[str]):
        self.channel = cols[_C_CHANNEL].strip()
        # nt_source(col1) — 인덱스 범위 방어(없으면 "").
        self.source = cols[_C_SOURCE].strip() if len(cols) > _C_SOURCE else ""
        self.medium = cols[_C_MEDIUM].strip()
        self.detail = cols[_C_DETAIL].strip()
        self.keyword = cols[_C_KEYWORD].strip()
        self.inflow = _num(cols[_C_INFLOW])
        self.pay = _num(cols[_C_PAY])
        self.amount = _num(cols[_C_AMOUNT])
        # +14일 기여도추정(보조) — 인덱스 범위 방어(없으면 0).
        self.pay14 = _num(cols[_C_PAY14]) if len(cols) > _C_PAY14 else 0.0
        self.amount14 = _num(cols[_C_AMOUNT14]) if len(cols) > _C_AMOUNT14 else 0.0


def _parse_rows(raw: str) -> list[_Row]:
    """raw → 데이터행만 _Row 리스트(헤더·깨진 줄 자동 제외)."""
    rows: list[_Row] = []
    for line in (raw or "").splitlines():
        if "\t" not in line:
            continue
        cols = line.split("\t")
        if len(cols) < _MIN_COLS:
            continue
        if cols[_C_CHANNEL].strip() not in _CHANNELS:
            continue
        if not _is_num(cols[_C_INFLOW]):
            continue
        rows.append(_Row(cols))
    return rows


def _aggregate(rows: list[_Row], key_fn, key_name: str, *, with_tag: bool = False) -> list[dict]:
    """key_fn(row) 로 그룹 합산 → [{key_name, inflow, pay, rate, amount[, tag]}], rate 내림차순."""
    groups: dict[str, dict] = {}
    order: list[str] = []
    for r in rows:
        k = key_fn(r)
        if k not in groups:
            groups[k] = {"inflow": 0.0, "pay": 0.0, "amount": 0.0,
                         "pay14": 0.0, "amount14": 0.0}
            order.append(k)
        g = groups[k]
        g["inflow"] += r.inflow
        g["pay"] += r.pay
        g["amount"] += r.amount
        g["pay14"] += r.pay14
        g["amount14"] += r.amount14
    out: list[dict] = []
    for k in order:
        g = groups[k]
        rate = _rate(g["pay"], g["inflow"])
        item = {key_name: k, "inflow": g["inflow"], "pay": g["pay"],
                "rate": rate, "amount": g["amount"],
                "pay14": g["pay14"], "amount14": g["amount14"]}
        if with_tag:
            item["tag"] = _perf_tag(rate)
        out.append(item)
    out.sort(key=lambda d: d["rate"], reverse=True)
    return out


def _build_warnings(device_rows: list[_Row]) -> list[str]:
    warnings: list[str] = []

    # ① nt_medium 대소문자 혼용
    spellings: dict[str, set] = {}
    for r in device_rows:
        if r.medium:
            spellings.setdefault(r.medium.lower(), set()).add(r.medium)
    if any(len(v) >= 2 for v in spellings.values()):
        warnings.append("nt_medium 대소문자 혼용(예: REVU/revu) — 분리 집계됨")

    # ② nt_keyword 빈 비율
    empty_kw = sum(1 for r in device_rows if r.keyword in ("", "-"))
    if empty_kw > 0:
        warnings.append(f"nt_keyword {empty_kw}개 비어있음 — 추적 약화")

    # ③ nt_detail 형식 혼용
    has_new = any(_is_new_detail(r.detail) for r in device_rows if r.detail)
    has_old = any((not _is_new_detail(r.detail)) for r in device_rows if r.detail)
    if has_new and has_old:
        warnings.append("nt_detail 형식 2종 혼용 — 규칙 통일 필요")

    return warnings


def parse_smartstore_table(raw: str) -> dict:
    """스마트스토어 마케팅분석 표 raw → 집계 dict.

    반환 키: summary, by_medium, by_campaign, by_product, warnings.
    """
    rows = _parse_rows(raw)
    total_rows = [r for r in rows if r.channel == _DEVICE_ALL]
    device_rows = [r for r in rows if r.channel in _DEVICE_SPLIT]

    # summary — '전체'행 합(없으면 기기행 합).
    src = total_rows if total_rows else device_rows
    s_inflow = sum(r.inflow for r in src)
    s_pay = sum(r.pay for r in src)
    s_amount = sum(r.amount for r in src)
    summary = {
        "inflow": s_inflow,
        "pay": s_pay,
        "pay_rate": _rate(s_pay, s_inflow),
        "amount": s_amount,
        "pay14": sum(r.pay14 for r in src),
        "amount14": sum(r.amount14 for r in src),
    }

    by_medium = _aggregate(device_rows, lambda r: r.medium.lower(), "medium")
    by_campaign = _aggregate(
        device_rows, lambda r: r.detail, "campaign", with_tag=True)

    # by_product — 신규형 행이 있을 때만(신규형=토큰[1], 구형=substring).
    has_new = any(_is_new_detail(r.detail) for r in device_rows if r.detail)
    if has_new:
        def _product_key(r: _Row) -> str:
            if _is_new_detail(r.detail):
                return r.detail.split("_")[1]
            return _detect_product(r.detail)
        by_product = _aggregate(device_rows, _product_key, "product")
    else:
        by_product = []

    # 파싱된 전체 데이터행(전체+기기, 붙여넣기 순서) — 엑셀 시트1 재현용.
    parsed_rows = [
        {
            "channel": r.channel, "source": r.source, "medium": r.medium,
            "detail": r.detail, "keyword": r.keyword, "inflow": r.inflow,
            "pay": r.pay, "pay_rate": _rate(r.pay, r.inflow), "amount": r.amount,
            "pay14": r.pay14, "amount14": r.amount14,
        }
        for r in rows
    ]

    return {
        "summary": summary,
        "by_medium": by_medium,
        "by_campaign": by_campaign,
        "by_product": by_product,
        "warnings": _build_warnings(device_rows),
        "rows": parsed_rows,
    }


# ── 엑셀(xlsx) 내보내기 — 2시트(원본붙여넣기 + 분석결과) ─────────────────────
# 결제율은 ★두 칸(표시=소수1자리 / 정확=풀소수)으로 — 엑셀 재검산 시 반올림 오차 방지.
# ★결제율은 처음부터 '비율'로 저장(rate/100)하고 % 서식으로 보여준다(엑셀에서 바로 % 인식).
_FMT_INT = "#,##0"
_FMT_RATE_SHOW = "0.0%"
_FMT_RATE_EXACT = "0.0000%"


def _rate_show(rate: float) -> float:
    """표시용 결제율 비율 — 소수1자리 반올림 후 /100(예: 7.13 → 0.071)."""
    return round(rate, 1) / 100.0


def _rate_exact(rate: float) -> float:
    """정확 결제율 비율 — 풀소수 /100(예: 7.129337 → 0.07129337)."""
    return rate / 100.0


def build_analytics_xlsx(result: dict) -> bytes:
    """집계 result → xlsx bytes(시트: 원본붙여넣기 + 분석결과). 순수 함수.

    결제율은 비율(rate/100)로 저장 + % 서식. 분석결과 시트는 섹션/헤더 강조·정렬·
    테두리·열너비 자동·우수/저조 색으로 가독성을 높인다."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    # ── 서식 상수(openpyxl 지연 import 유지 위해 함수 내부 정의) ──
    head_fill = PatternFill("solid", fgColor="305496")   # 진한 남색
    head_font = Font(bold=True, color="FFFFFF")
    sect_fill = PatternFill("solid", fgColor="D9E1F2")    # 연한 남색
    sect_font = Font(bold=True, size=12, color="1F3864")
    good_font = Font(bold=True, color="006100")           # 우수 초록
    bad_font = Font(bold=True, color="9C0006")            # 저조 빨강
    _side = Side(style="thin", color="BFBFBF")
    thin = Border(left=_side, right=_side, top=_side, bottom=_side)
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")
    left = Alignment(horizontal="left")

    def _header_row(ws, values):
        ws.append(list(values))
        r = ws.max_row
        for col in range(1, len(values) + 1):
            c = ws.cell(row=r, column=col)
            c.fill, c.font, c.alignment, c.border = head_fill, head_font, center, thin
        return r

    def _section_row(ws, title, ncols):
        ws.append([title])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(row=r, column=1)
        c.fill, c.font = sect_fill, sect_font
        return r

    def _data_row(ws, values, *, num_cols=(), fmts=None):
        ws.append(list(values))
        r = ws.max_row
        for col in range(1, len(values) + 1):
            c = ws.cell(row=r, column=col)
            c.border = thin
            c.alignment = right if col in num_cols else left
        for col, fmt in (fmts or {}).items():
            ws.cell(row=r, column=col).number_format = fmt
        return r

    def _autosize(ws):
        widths: dict[str, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                letter = cell.column_letter
                widths[letter] = max(widths.get(letter, 0), len(str(cell.value)))
        for letter, w in widths.items():
            ws.column_dimensions[letter].width = max(10, min(40, w * 1.3))

    wb = Workbook()

    # ── 시트1: 원본붙여넣기(파싱된 전체 데이터행) ──
    ws1 = wb.active
    ws1.title = "원본붙여넣기"
    _header_row(ws1, [
        "채널속성", "nt_source", "nt_medium", "nt_detail", "nt_keyword",
        "유입", "결제", "결제율_표시", "결제율_정확", "결제금액",
        "결제수_14일", "결제금액_14일",
    ])
    row_num_cols = (6, 7, 8, 9, 10, 11, 12)
    row_fmt = {6: _FMT_INT, 7: _FMT_INT, 8: _FMT_RATE_SHOW,
               9: _FMT_RATE_EXACT, 10: _FMT_INT, 11: _FMT_INT, 12: _FMT_INT}
    for r in result.get("rows", []):
        _data_row(ws1, [
            r["channel"], r["source"], r["medium"], r["detail"], r["keyword"],
            r["inflow"], r["pay"], _rate_show(r["pay_rate"]), _rate_exact(r["pay_rate"]),
            r["amount"], r["pay14"], r["amount14"],
        ], num_cols=row_num_cols, fmts=row_fmt)
    ws1.freeze_panes = "A2"
    _autosize(ws1)

    # ── 시트2: 분석결과(섹션 세로 stack) ──
    ws2 = wb.create_sheet("분석결과")
    s = result["summary"]

    _section_row(ws2, "[요약]", 2)
    _data_row(ws2, ["총 유입", s["inflow"]], num_cols=(2,), fmts={2: _FMT_INT})
    _data_row(ws2, ["총 결제", s["pay"]], num_cols=(2,), fmts={2: _FMT_INT})
    _data_row(ws2, ["결제율_표시", _rate_show(s["pay_rate"])],
              num_cols=(2,), fmts={2: _FMT_RATE_SHOW})
    _data_row(ws2, ["결제율_정확", _rate_exact(s["pay_rate"])],
              num_cols=(2,), fmts={2: _FMT_RATE_EXACT})
    _data_row(ws2, ["결제금액", s["amount"]], num_cols=(2,), fmts={2: _FMT_INT})
    _data_row(ws2, ["결제수(+14일)", s.get("pay14", 0)], num_cols=(2,), fmts={2: _FMT_INT})
    _data_row(ws2, ["결제금액(+14일)", s.get("amount14", 0)],
              num_cols=(2,), fmts={2: _FMT_INT})
    for rr in range(ws2.max_row - 6, ws2.max_row + 1):    # 항목명(A열) bold(7개)
        ws2.cell(row=rr, column=1).font = Font(bold=True)

    ws2.append([])
    _section_row(ws2, "[데이터 품질 경고]", 2)
    for w in (result.get("warnings") or ["규칙 위반 없음"]):
        _data_row(ws2, [w])

    # 매체/캠페인/제품 공통 숫자열·포맷(+14일 결제수·결제금액 2열 포함, 결제율14는 없음).
    agg_num = (2, 3, 4, 5, 6, 7, 8)
    agg_fmt = {2: _FMT_INT, 3: _FMT_INT, 4: _FMT_RATE_SHOW,
               5: _FMT_RATE_EXACT, 6: _FMT_INT, 7: _FMT_INT, 8: _FMT_INT}
    agg_head = ["유입", "결제", "결제율_표시", "결제율_정확", "결제금액",
                "결제수_14일", "결제금액_14일"]

    def _agg_vals(d, name_key):
        return [
            d[name_key], d["inflow"], d["pay"], _rate_show(d["rate"]),
            _rate_exact(d["rate"]), d["amount"],
            d.get("pay14", 0), d.get("amount14", 0),
        ]

    ws2.append([])
    _section_row(ws2, "[매체 통합 집계]", 8)
    _header_row(ws2, ["매체", *agg_head])
    for m in result.get("by_medium", []):
        _data_row(ws2, _agg_vals(m, "medium"), num_cols=agg_num, fmts=agg_fmt)

    ws2.append([])
    _section_row(ws2, "[캠페인별 성과]", 9)
    _header_row(ws2, ["캠페인", *agg_head, "태그"])
    _tag_kr = {"good": "우수", "bad": "저조", "mid": ""}
    for c in result.get("by_campaign", []):
        tag = _tag_kr.get(c.get("tag", ""), "")
        r = _data_row(ws2, [*_agg_vals(c, "campaign"), tag],
                      num_cols=agg_num, fmts=agg_fmt)
        tag_cell = ws2.cell(row=r, column=9)
        tag_cell.alignment = center
        if tag == "우수":
            tag_cell.font = good_font
        elif tag == "저조":
            tag_cell.font = bad_font

    if result.get("by_product"):
        ws2.append([])
        _section_row(ws2, "[제품별 성과]", 8)
        _header_row(ws2, ["제품", *agg_head])
        for p in result["by_product"]:
            _data_row(ws2, _agg_vals(p, "product"), num_cols=agg_num, fmts=agg_fmt)

    _autosize(ws2)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
