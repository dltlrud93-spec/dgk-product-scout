"""
teamp_mode.py — 체험단 타겟 선정 모드 코어.

제품 키워드(사용자 입력) → 차종 수확(꼬리물기) → (차종×제품) 행마다
  검색량(합산) + 블로그 문서수 + 비율 → 황금/해볼만/포화 3분류.

블로그 문서수:
  네이버 블로그 검색 API(GET https://openapi.naver.com/v1/search/blog.json)
  "{차종 표시명} {product}" 쿼리 → 응답 total 값.
  ★ total 은 근사치이고 블로그 노출 난이도(블로그 지수)와 다름 — 우선순위 신호이지 노출 보증 아님.

비율 = 문서수 ÷ 검색량. 낮을수록 기회(글이 수요보다 적음).
검색량 0(원래 '<10') 차종은 volume 필터로 먼저 제외 → 블로그 API 미호출.
분류 임계는 config.TEAMP_RATIO_GOLD / TEAMP_RATIO_OK.

429 대응: call_delay → fetch_blog_count 호출 전 지연. 429 응답은 지수백오프 재시도.
부분 실패: fetch_teamp_rows_partial 에서 항목별 예외 격리 → 실패 항목만 failed 리스트.
"""

from __future__ import annotations

import calendar
import datetime
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from typing import Callable, Optional

import requests

import config
from src.core.search_volume import dedupe_relkeywords, member_volume


class BlogFetchError(RuntimeError):
    """블로그 문서수 조회 실패 (429 재시도 소진 또는 기타 오류)."""


@dataclass
class TeampRow:
    canonical: str    # 차종 정규명 (인식 사전 기준)
    product: str      # 제품 키워드 (사용자 입력)
    volume: int       # 검색량(합산) — harvest_models 의 agg["volume"] 그대로
    doc_count: int    # 블로그 문서수 (네이버 블로그 검색 total)
    ratio: float      # doc_count / volume
    grade: str        # "🟡 황금" / "🟢 해볼만" / "🔴 포화/후순위"


def _query_name(canonical: str) -> str:
    """블로그 검색 쿼리용 차종 표시명.

    인식 사전의 정규명에 포함되는 '(세대미상)' 등 괄호 표기를 제거하고
    실제 사람들이 검색하는 차종명만 남긴다.

    예) "모닝(세대미상)" → "모닝"
        "아반떼(세대미상)" → "아반떼"
        "아반떼CN7" → "아반떼CN7"   (괄호 없으면 그대로)
        "레이" → "레이"

    ★ 이유: canonical="모닝(세대미상)"을 그대로 query에 쓰면 "모닝(세대미상) 에어컨필터"로
      검색되어 블로그 결과가 극소 (~630) → 비율 0.13(황금 오분류). 실제 검색어인
      "모닝 에어컨필터"를 쓰면 ~18,850개로 올바른 분류(포화)가 된다.
    """
    return re.sub(r"\(.*?\)", "", canonical).strip()


def classify_ratio(
    ratio: float,
    *,
    gold: float = config.TEAMP_RATIO_GOLD,
    ok: float = config.TEAMP_RATIO_OK,
) -> str:
    """비율 → 3분류.

    · ratio < gold  → 🟡 황금   (글이 수요보다 적어 노릴 자리)
    · gold ≤ ratio ≤ ok → 🟢 해볼만
    · ratio > ok    → 🔴 포화/후순위
    """
    if ratio < gold:
        return "🟡 황금"
    if ratio <= ok:
        return "🟢 해볼만"
    return "🔴 포화/후순위"


def now_kst_str() -> str:
    """현재 시각을 KST 'YYYY-MM-DD HH:MM' 문자열로 반환(수집 시각 표기용).

    zoneinfo tz 데이터가 없는 환경(윈도우 등)에서도 안전하도록 고정 오프셋(+9h) 사용.
    """
    kst = datetime.timezone(datetime.timedelta(hours=9))
    return datetime.datetime.now(kst).strftime("%Y-%m-%d %H:%M")


def opportunity_score(volume: int, doc_count: int, k: float = config.TEAMP_EV_K) -> float:
    """기회 점수(EV) = 검색량 × 상위노출 성공확률 근사.

    성공확률 근사 = k / (k + doc_count) — 문서수가 많을수록(경쟁 깊을수록) 감소.
    ★ 순위 비교 전용 지표이지 유입 예측치가 아니다. k 는 상위노출 진입 난이도 상수.
    """
    return volume * k / (k + doc_count)


def split_priority_groups(rows) -> tuple[list, list, list]:
    """등급 문자열 앞글자로 3그룹 분리 — (gold, ok, saturated). 원소 순서 보존.

    · grade 가 "🟡"로 시작 → gold
    · "🟢"로 시작 → ok
    · 그 외("🔴" 등) → saturated
    """
    gold: list = []
    ok: list = []
    saturated: list = []
    for r in rows:
        if r.grade.startswith("🟡"):
            gold.append(r)
        elif r.grade.startswith("🟢"):
            ok.append(r)
        else:
            saturated.append(r)
    return gold, ok, saturated


def sort_rows_for_display(rows, sort_label) -> list:
    """정렬 라벨에 따라 rows 를 정렬한 새 리스트 반환(원본 불변).

    · "기회 점수순 (추천)": opportunity_score 내림차순 → 동률 시 volume 내림차순 → keyword 오름차순
    · "검색량 내림차순": volume 내림차순
    · "검색량↑ + 최근글↓ (숨은 기회)": volume 내림차순, 동률 시 recent_3m_docs 오름차순(None=inf)
    · 그 외("비율 오름차순 (황금 위)" 등): ratio 오름차순
    """
    if sort_label == "기회 점수순 (추천)":
        return sorted(
            rows,
            key=lambda r: (-opportunity_score(r.volume, r.doc_count), -r.volume, r.keyword),
        )
    if sort_label == "검색량 내림차순":
        return sorted(rows, key=lambda r: r.volume, reverse=True)
    if sort_label == "검색량↑ + 최근글↓ (숨은 기회)":
        return sorted(
            rows,
            key=lambda r: (
                -r.volume,
                r.recent_3m_docs if r.recent_3m_docs is not None else float("inf"),
            ),
        )
    return sorted(rows, key=lambda r: r.ratio)


def _cutoff_date(months: int) -> datetime.date:
    """오늘 기준 months 개월 전 날짜. 말일 clamp 처리(예: 3/31 → 3개월 전 = 12/31)."""
    today = datetime.date.today()
    month = today.month - months
    year = today.year
    while month <= 0:
        month += 12
        year -= 1
    last_day = calendar.monthrange(year, month)[1]
    day = min(today.day, last_day)
    return datetime.date(year, month, day)


def _parse_postdate(postdate: str) -> Optional[datetime.date]:
    """'20240115' → datetime.date. 파싱 불가 시 None."""
    try:
        return datetime.date(int(postdate[:4]), int(postdate[4:6]), int(postdate[6:8]))
    except (ValueError, IndexError, TypeError):
        return None


def fetch_recent_blog_count(
    query: str,
    client_id: str,
    client_secret: str,
    *,
    months: int = config.TEAMP_RECENT_MONTHS,
    http_get: Optional[Callable] = None,
    sleep_fn: Optional[Callable[[float], None]] = None,
    max_retries: int = config.NAVER_BLOG_MAX_RETRIES,
    backoff_seconds: float = config.NAVER_BLOG_BACKOFF_SECONDS,
    call_delay: float = config.NAVER_BLOG_CALL_DELAY,
) -> int:
    """네이버 블로그 검색(sort=date, display=100) → 최근 months 개월 이내 글 수 추정.

    최신순 상위 NAVER_BLOG_SEARCH_RECENT_DISPLAY 건의 postdate 를 파싱해 카운트.
    ★ 추정치: 상위 100건 기준이라 전수가 아님. 총 문서수 ≤ 100 이면 정확도 높음.
    429/백오프/재시도는 fetch_blog_count 와 동일 패턴.
    """
    _get = http_get if http_get is not None else requests.get
    _sleep = sleep_fn if sleep_fn is not None else time.sleep

    _sleep(call_delay)

    cutoff = _cutoff_date(months)

    for attempt in range(max_retries + 1):
        resp = _get(
            config.NAVER_BLOG_SEARCH_URL,
            params={
                "query": query,
                "display": config.NAVER_BLOG_SEARCH_RECENT_DISPLAY,
                "sort": "date",
            },
            headers={
                "X-Naver-Client-Id": client_id,
                "X-Naver-Client-Secret": client_secret,
            },
            timeout=10,
        )
        if resp.status_code == 429:
            if attempt < max_retries:
                retry_after = resp.headers.get("Retry-After")
                try:
                    wait = float(retry_after) if retry_after else backoff_seconds * (2 ** attempt)
                except ValueError:
                    wait = backoff_seconds * (2 ** attempt)
                _sleep(wait)
                continue
            raise BlogFetchError(
                f"429 Too Many Requests — {max_retries}회 재시도 소진: {query!r}"
            )
        resp.raise_for_status()
        items = resp.json().get("items", [])
        return sum(
            1 for item in items
            if (_parse_postdate(item.get("postdate", "")) or datetime.date.min) >= cutoff
        )

    raise BlogFetchError(f"재시도 소진: {query!r}")  # 실질적 도달 불가


def fetch_blog_count(
    query: str,
    client_id: str,
    client_secret: str,
    *,
    http_get: Optional[Callable] = None,
    sleep_fn: Optional[Callable[[float], None]] = None,
    max_retries: int = config.NAVER_BLOG_MAX_RETRIES,
    backoff_seconds: float = config.NAVER_BLOG_BACKOFF_SECONDS,
    call_delay: float = config.NAVER_BLOG_CALL_DELAY,
) -> int:
    """네이버 블로그 검색 API → 문서수(total).

    호출 전 call_delay 초 대기(rate limit 완충).
    429 응답: Retry-After 헤더 우선, 없으면 지수백오프(1→2→4초). 최대 max_retries 회.
    재시도 소진 시 BlogFetchError 발생.

    ★ total 은 근사치(네이버 내부 추정). 블로그 지수(상위 글 품질)는 미반영.
    """
    _get = http_get if http_get is not None else requests.get
    _sleep = sleep_fn if sleep_fn is not None else time.sleep

    _sleep(call_delay)

    for attempt in range(max_retries + 1):
        resp = _get(
            config.NAVER_BLOG_SEARCH_URL,
            params={"query": query, "display": config.NAVER_BLOG_SEARCH_DISPLAY},
            headers={
                "X-Naver-Client-Id": client_id,
                "X-Naver-Client-Secret": client_secret,
            },
            timeout=10,
        )
        if resp.status_code == 429:
            if attempt < max_retries:
                retry_after = resp.headers.get("Retry-After")
                try:
                    wait = float(retry_after) if retry_after else backoff_seconds * (2 ** attempt)
                except ValueError:
                    wait = backoff_seconds * (2 ** attempt)
                _sleep(wait)
                continue
            raise BlogFetchError(
                f"429 Too Many Requests — {max_retries}회 재시도 소진: {query!r}"
            )
        resp.raise_for_status()
        return int(resp.json().get("total", 0))

    raise BlogFetchError(f"재시도 소진: {query!r}")  # 실질적 도달 불가(루프 구조상)


_HTML_TAG_RE = re.compile(r"<[^>]+>")


def _strip_html(text: str) -> str:
    """블로그 검색 title 의 <b></b> 강조 태그·HTML 엔티티(&amp; 등)를 제거해 순수 텍스트로."""
    import html

    return html.unescape(_HTML_TAG_RE.sub("", text or "")).strip()


def fetch_blog_titles(
    query: str,
    client_id: str,
    client_secret: str,
    *,
    display: int = config.NAVER_BLOG_TITLE_DISPLAY,
    http_get: Optional[Callable] = None,
    sleep_fn: Optional[Callable[[float], None]] = None,
    max_retries: int = config.NAVER_BLOG_MAX_RETRIES,
    backoff_seconds: float = config.NAVER_BLOG_BACKOFF_SECONDS,
    call_delay: float = config.NAVER_BLOG_CALL_DELAY,
) -> list[str]:
    """네이버 블로그 검색(정확도순, display=N) → 글 제목 리스트(HTML 태그 제거).

    체험단 양식 키워드 추천의 '블로그 제목 기반' 보완용. 검색량 무관하게 실제 글
    제목을 받아 표현을 확보한다. 429/백오프/재시도는 fetch_blog_count 와 동일 패턴.
    """
    _get = http_get if http_get is not None else requests.get
    _sleep = sleep_fn if sleep_fn is not None else time.sleep

    _sleep(call_delay)

    for attempt in range(max_retries + 1):
        resp = _get(
            config.NAVER_BLOG_SEARCH_URL,
            params={"query": query, "display": display},
            headers={
                "X-Naver-Client-Id": client_id,
                "X-Naver-Client-Secret": client_secret,
            },
            timeout=10,
        )
        if resp.status_code == 429:
            if attempt < max_retries:
                retry_after = resp.headers.get("Retry-After")
                try:
                    wait = float(retry_after) if retry_after else backoff_seconds * (2 ** attempt)
                except ValueError:
                    wait = backoff_seconds * (2 ** attempt)
                _sleep(wait)
                continue
            raise BlogFetchError(
                f"429 Too Many Requests — {max_retries}회 재시도 소진: {query!r}"
            )
        resp.raise_for_status()
        items = resp.json().get("items", [])
        titles = [_strip_html(item.get("title", "")) for item in items]
        return [t for t in titles if t]

    raise BlogFetchError(f"재시도 소진: {query!r}")  # 실질적 도달 불가(루프 구조상)


def fetch_teamp_rows_partial(
    valid_items: list[tuple[str, str, int]],
    blog_fn: Callable[[str], int],
    max_workers: int = config.NAVER_BLOG_MAX_WORKERS,
    on_progress: Optional[Callable[[int, int], None]] = None,
) -> tuple[list[TeampRow], list[tuple[str, str, int]]]:
    """(canonical, ptype, volume) 리스트 → (성공 TeampRow 목록, 실패 항목 목록).

    blog_fn 이 예외를 발생시키면 해당 항목을 failed 에 수집하고 나머지는 계속 처리.
    on_progress(done, total) 콜백 — 항목 완료마다 호출(선택).
    반환된 rows 는 ratio 오름차순 정렬.
    """
    if not valid_items:
        return [], []

    total = len(valid_items)
    rows: list[TeampRow] = []
    failed: list[tuple[str, str, int]] = []
    done = 0

    with ThreadPoolExecutor(max_workers=min(max_workers, total)) as executor:
        future_map = {
            executor.submit(blog_fn, f"{_query_name(c)} {p}"): (c, p, v)
            for c, p, v in valid_items
        }
        for future in as_completed(future_map):
            c, p, v = future_map[future]
            try:
                doc_count = future.result()
                ratio = doc_count / v
                rows.append(TeampRow(c, p, v, doc_count, ratio, classify_ratio(ratio)))
            except Exception:
                failed.append((c, p, v))
            done += 1
            if on_progress:
                on_progress(done, total)

    rows.sort(key=lambda r: r.ratio)
    return rows, failed


def build_teamp_rows(
    agg: dict,
    client_id: str,
    client_secret: str,
    *,
    blog_fetch_fn: Optional[Callable[[str], int]] = None,
    max_workers: int = 1,
) -> list[TeampRow]:
    """
    harvest_models 결과(agg) → TeampRow 리스트.

    agg 구조: {(canonical, ptype): {"volume": int, "members": int, ...}}
    harvest_models 호출 시 part_seeds = {product_kw: [product_kw]} 로 구성하면
    ptype == product_keyword 가 된다.

    처리 순서:
    1. volume == 0(원래 '<10') 행 먼저 제외 — 블로그 API 미호출.
    2. 남은 행: _query_name(canonical)으로 쿼리 구성 ("모닝(세대미상)" → "모닝").
    3. max_workers > 1 이면 ThreadPoolExecutor 병렬, 아니면 순차 실행.
    4. 비율 오름차순 정렬(황금이 위).
    """
    _blog = blog_fetch_fn or (lambda q: fetch_blog_count(q, client_id, client_secret))

    # Step 1: volume=0 먼저 필터 (블로그 API 호출 전)
    valid = [
        (canonical, ptype, a["volume"])
        for (canonical, ptype), a in agg.items()
        if a["volume"] > 0
    ]

    if not valid:
        return []

    rows: list[TeampRow] = []

    if max_workers > 1:
        # 병렬 실행 (Streamlit 앱 경로는 render_teamp 에서 직접 처리 — 여기는 직접 호출용)
        with ThreadPoolExecutor(max_workers=min(max_workers, len(valid))) as executor:
            future_map = {
                executor.submit(_blog, f"{_query_name(c)} {p}"): (c, p, v)
                for c, p, v in valid
            }
            for future in as_completed(future_map):
                c, p, v = future_map[future]
                doc_count = future.result()
                ratio = doc_count / v
                rows.append(TeampRow(c, p, v, doc_count, ratio, classify_ratio(ratio)))
    else:
        for c, p, v in valid:
            doc_count = _blog(f"{_query_name(c)} {p}")
            ratio = doc_count / v
            rows.append(TeampRow(c, p, v, doc_count, ratio, classify_ratio(ratio)))

    rows.sort(key=lambda r: r.ratio)
    return rows


def top_gold_rows(rows: list[TeampRow], n: int = 10) -> list[TeampRow]:
    """🟡 황금 행 중 검색량 높은 순 상위 n개."""
    gold = [r for r in rows if r.grade == "🟡 황금"]
    gold.sort(key=lambda r: r.volume, reverse=True)
    return gold[:n]


def top_gold_rows_by_ratio(rows: list[TeampRow], n: int = 10) -> list[TeampRow]:
    """🟡 황금 행 중 비율 낮은 순 상위 n개."""
    gold = [r for r in rows if r.grade == "🟡 황금"]
    gold.sort(key=lambda r: r.ratio)
    return gold[:n]


# ═══════════════════════════════════════════════════════════════════════════════
# 키워드 단위 체험단 모드 — 합산 없음, 개별 키워드가 단위
# ═══════════════════════════════════════════════════════════════════════════════
#
# 핵심 원칙:
#   · 단위 = 개별 연관 키워드(예: "셀토스에어컨필터", "셀토스자동차에어컨필터" 각각 별도 행).
#   · 검색량 = 그 키워드의 monthlyPcQcCnt + monthlyMobileQcCnt (합산 아님).
#   · 블로그 문서수 = keyword 원문 그대로 네이버 블로그 검색 → total.
#     ★ 검색량을 잰 키워드 = 문서수를 재는 쿼리 — 100% 동일 문자열.
#   · 차종 인식은 표시 전용(car_model 컬럼) — 필터·그룹핑·합산에 사용 금지.


@dataclass
class TeampKwRow:
    keyword: str       # 연관 키워드 원문 (예: "셀토스에어컨필터")
    car_model: str     # 차종 표시용 (차종 인식 사전 → canonical. 미인식이면 "")
    volume: int        # 개별 키워드 검색량 (monthlyPcQcCnt + monthlyMobileQcCnt)
    doc_count: int     # 블로그 문서수 (keyword 원문 그대로 블로그 검색 → total)
    ratio: float       # doc_count / volume
    grade: str         # "🟡 황금" / "🟢 해볼만" / "🔴 포화/후순위"
    recent_3m_docs: Optional[int] = None  # 최근 N개월 이내 블로그 글 수 추정치. None = 조회 안 함("—")


def harvest_teamp_kw_items(
    adapter,
    products: list[str],
    index,
) -> list[tuple[str, str, int]]:
    """제품 키워드들로 연관 키워드 수확 → 제품명 포함 + volume>0 필터.

    반환: [(keyword, car_model_display, volume), ...] volume 내림차순.
    ★ index.recognize 는 표시용 car_model 컬럼만 — 필터·그룹핑 미사용.
    ★ 블로그 검색 쿼리 = keyword 원문 그대로 (차종명 정제·공백 삽입·괄호 제거 없음).
    """
    product_terms = [p.strip().lower() for p in products if p.strip()]

    def _contains_product(kw: str) -> bool:
        kw_l = kw.lower()
        return any(term in kw_l for term in product_terms)

    all_kws: dict[str, dict] = {}  # relKeyword → raw row (시드 간 dedupe)
    for i, seed in enumerate(products):
        seed = seed.strip()
        if not seed:
            continue
        if i > 0:
            adapter._sleep(adapter.rate_limit_seconds)
        rows = adapter._request_keywordstool([seed])
        for kw, row in dedupe_relkeywords(rows).items():
            if kw not in all_kws:
                all_kws[kw] = row

    result: list[tuple[str, str, int]] = []
    for kw, row in all_kws.items():
        if not _contains_product(kw):
            continue
        vol = member_volume(row)
        if vol == 0:
            continue
        rec = index.recognize(kw)
        car_model = rec.canonical if rec.recognized else ""
        result.append((kw, car_model, vol))

    result.sort(key=lambda x: x[2], reverse=True)
    return result


def fetch_teamp_kw_rows_partial(
    kw_items: list[tuple[str, str, int]],
    blog_fn: Callable[[str], int],
    max_workers: int = config.NAVER_BLOG_MAX_WORKERS,
    on_progress: Optional[Callable[[int, int], None]] = None,
) -> tuple[list[TeampKwRow], list[tuple[str, str, int]]]:
    """(keyword, car_model, volume) 리스트 → (성공 TeampKwRow 목록, 실패 항목 목록).

    ★ blog_fn 에 keyword 원문 그대로 전달 — _query_name 등 변환 없음.
    blog_fn 예외 발생 시 해당 항목 failed 수집 후 계속 처리(부분 실패 허용).
    반환된 rows 는 ratio 오름차순 정렬.
    """
    if not kw_items:
        return [], []

    total = len(kw_items)
    rows: list[TeampKwRow] = []
    failed: list[tuple[str, str, int]] = []
    done = 0

    with ThreadPoolExecutor(max_workers=min(max_workers, total)) as executor:
        future_map = {
            executor.submit(blog_fn, keyword): (keyword, car_model, volume)
            for keyword, car_model, volume in kw_items
        }
        for future in as_completed(future_map):
            keyword, car_model, volume = future_map[future]
            try:
                doc_count = future.result()
                ratio = doc_count / volume
                rows.append(TeampKwRow(keyword, car_model, volume, doc_count, ratio, classify_ratio(ratio)))
            except Exception:
                failed.append((keyword, car_model, volume))
            done += 1
            if on_progress:
                on_progress(done, total)

    rows.sort(key=lambda r: r.ratio)
    return rows, failed


def top_gold_kw_rows(rows: list[TeampKwRow], n: int = 10) -> list[TeampKwRow]:
    """🟡 황금 키워드 중 검색량 높은 순 상위 n개."""
    gold = [r for r in rows if r.grade == "🟡 황금"]
    gold.sort(key=lambda r: r.volume, reverse=True)
    return gold[:n]


def top_gold_kw_rows_by_ratio(rows: list[TeampKwRow], n: int = 10) -> list[TeampKwRow]:
    """🟡 황금 키워드 중 비율 낮은 순 상위 n개."""
    gold = [r for r in rows if r.grade == "🟡 황금"]
    gold.sort(key=lambda r: r.ratio)
    return gold[:n]


def format_recent_3m(val: Optional[int]) -> str:
    """최근3개월 → 신호등 라벨 (표시 전용).

    None(미조회·실패) → '—'. 숫자 비교는 None 가드 이후에만 수행.
    임계: config.TEAMP_RECENT_HOT(100) / TEAMP_RECENT_BUSY(30) / TEAMP_RECENT_GOOD(6).
    """
    if val is None:
        return "—"
    if val >= config.TEAMP_RECENT_HOT:
        return "🔴 100+ 비추천"
    if val >= config.TEAMP_RECENT_BUSY:
        return f"🟡 {val} 보통"
    if val >= config.TEAMP_RECENT_GOOD:
        return f"🟢 {val} 노려볼만"
    return f"🟢 {val} 최고"


def format_recent_ratio(recent_3m_docs: Optional[int], doc_count: Optional[int]) -> str:
    """최근비중(%) = 최근3개월 ÷ 전체문서 × 100 (표시 전용).

    '—' 반환 조건 (산술 전에 전부 가드):
    · recent_3m_docs None(미조회·실패)
    · recent_3m_docs ≥ 상한(NAVER_BLOG_SEARCH_RECENT_DISPLAY) — 분자가 과소(실제는 그 이상일 수 있음)
    · doc_count None 또는 0 — 분모 없음
    """
    if recent_3m_docs is None:
        return "—"
    if recent_3m_docs >= config.NAVER_BLOG_SEARCH_RECENT_DISPLAY:
        return "—"
    if not doc_count:
        return "—"
    pct = recent_3m_docs / doc_count * 100
    return f"{pct:.1f}%"


def fetch_recent_3m_docs_partial(
    rows: list[TeampKwRow],
    recent_blog_fn: Callable[[str], int],
    max_workers: int = config.NAVER_BLOG_MAX_WORKERS,
    on_progress: Optional[Callable[[int, int], None]] = None,
) -> list[TeampKwRow]:
    """rows 의 recent_3m_docs 를 채워 반환 (in-place 갱신 후 rows 반환).

    조회 대상:
    · 황금(🟡) + 해볼만(🟢) → 전체
    · 포화(🔴) 중 검색량 ≥ config.TEAMP_SATURATED_MIN_VOLUME → 숨은 기회 탐색
    조회 안 함 → recent_3m_docs 는 None 유지 (표시 "—").
    실패 항목: recent_3m_docs = None (부분 실패 허용, 전체 중단 없음).
    """
    target = [
        r for r in rows
        if r.grade != "🔴 포화/후순위" or r.volume >= config.TEAMP_SATURATED_MIN_VOLUME
    ]
    if not target:
        return rows

    total = len(target)
    done = 0
    result_map: dict[str, Optional[int]] = {}

    with ThreadPoolExecutor(max_workers=min(max_workers, total)) as executor:
        future_map = {
            executor.submit(recent_blog_fn, r.keyword): r.keyword
            for r in target
        }
        for future in as_completed(future_map):
            keyword = future_map[future]
            try:
                result_map[keyword] = future.result()
            except Exception:
                result_map[keyword] = None
            done += 1
            if on_progress:
                on_progress(done, total)

    for r in rows:
        if r.keyword in result_map:
            r.recent_3m_docs = result_map[r.keyword]

    return rows


# ── 탭 전환 시 사이드바 위젯 복원(순수) ───────────────────────────────────────
# 배경: 다른 화면으로 갔다 오면 사이드바 위젯이 evict 돼 기본값으로 리셋된다. 그러면
# 요청 서명(소스·제품·데이터제품·상한)이 바뀌어 캐시 미적중 → 블로그 문서수 등 API 재호출.
# 마지막 선택을 비위젯 키(_teamp_last_*)에 백업해두고, 위젯 렌더 전에 여기서 복원한다.

def restore_teamp_widgets(state, *, src_opts, jp_opts, sort_opts=None, top10_opts=None):
    """evict 된 사이드바 위젯 키를 마지막 백업값으로 복원(순수 — state 는 dict 유사).

    백업키(_teamp_last_*)가 있고 위젯키가 ★없을 때만, 그리고 ★유효 옵션일 때만 복원
    (옵션 변경 시 StreamlitAPIException 방지). 위젯키가 이미 있으면 사용자 선택이므로 불변.
    """
    last_kws = state.get("_teamp_last_keywords")
    if last_kws and "teamp_products" not in state:
        state["teamp_products"] = ", ".join(last_kws)
    last_src = state.get("_teamp_last_source")
    if last_src in src_opts and "teamp_source" not in state:
        state["teamp_source"] = last_src
    last_jp = state.get("_teamp_last_jp_product")
    if last_jp in jp_opts and "teamp_jp_product" not in state:
        state["teamp_jp_product"] = last_jp
    if sort_opts is not None:
        last_sort = state.get("_teamp_last_sort")
        if last_sort in sort_opts and "teamp_sort" not in state:
            state["teamp_sort"] = last_sort
    if top10_opts is not None:
        last_t10 = state.get("_teamp_last_top10_sort")
        if last_t10 in top10_opts and "teamp_top10_sort" not in state:
            state["teamp_top10_sort"] = last_t10


def normalize_teamp_cache(raw):
    """세션의 _teamp_results 를 {signature: 결과dict} 맵으로 정규화한다(순수).

    소스별 결과를 서명(signature)별로 따로 보관하기 위한 맵. 구버전(단일 결과 dict,
    'rows' 키 보유) 또는 None/비-dict 이면 빈 맵을 돌려준다(안전 초기화).

    ★ collected_at 키가 없는 구버전 엔트리도 그대로 허용한다(엔트리 내부는 손대지 않음).
      수집 시각은 읽는 쪽에서 entry.get("collected_at")→None 으로 방어한다."""
    if not isinstance(raw, dict) or "rows" in raw:
        return {}
    return raw


def build_teamp_xlsx(gold, ok, saturated) -> bytes:
    """순위 3그룹 → xlsx bytes. 시트 3개(1순위 황금 / 2순위 해볼만 / 3순위 포화).

    · 각 시트 행 순서 = 기회 점수 내림차순(화면 기본 정렬과 동일).
    · 컬럼: 키워드 | 기회 점수 | 검색량 | 문서수 | 비율 | 최근3개월 | 최근비중 | 차종.
    · 기회 점수 = round(opportunity_score) 정수 · 검색량/문서수 정수 · 비율 소수2자리 숫자.
    · 최근3개월/최근비중은 화면 표기 문자열(format_recent_3m/format_recent_ratio) 그대로.
    openpyxl 은 함수 내부에서 지연 import(선택적 의존성 최상단 import 금지 원칙)."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    header = ["키워드", "기회 점수", "검색량", "문서수", "비율", "최근3개월", "최근비중", "차종"]
    head_fill = PatternFill("solid", fgColor="305496")   # 진한 남색(성과분석 xlsx 와 동일 톤)
    head_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")

    sheets = [
        ("1순위 황금", gold),
        ("2순위 해볼만", ok),
        ("3순위 포화(비추천)", saturated),
    ]

    wb = Workbook()
    for i, (title, group_rows) in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = title
        ws.append(header)
        for col in range(1, len(header) + 1):
            c = ws.cell(row=1, column=col)
            c.fill, c.font, c.alignment = head_fill, head_font, center
        for r in sort_rows_for_display(group_rows, "기회 점수순 (추천)"):
            ws.append([
                r.keyword,
                int(round(opportunity_score(r.volume, r.doc_count))),
                int(r.volume),
                int(r.doc_count),
                round(r.ratio, 2),
                format_recent_3m(r.recent_3m_docs),
                format_recent_ratio(r.recent_3m_docs, r.doc_count),
                r.car_model,
            ])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
