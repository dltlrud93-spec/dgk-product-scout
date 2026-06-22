"""
test_campaign_analytics.py — 스마트스토어 마케팅분석 표 파싱·집계 검증(순수 함수).
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from src.core.campaign_analytics import (
    _is_num,
    _num,
    build_analytics_xlsx,
    parse_smartstore_table,
)


def _row(channel, medium, detail, keyword, inflow, pay, amount,
         pay14="0", amount14="0"):
    """16컬럼 데이터행(인덱스 0/2/3/4/6/9/11 + +14일 13/15)."""
    cols = ["-"] * 16
    cols[0] = channel
    cols[2] = medium
    cols[3] = detail
    cols[4] = keyword
    cols[6] = inflow
    cols[9] = pay
    cols[11] = amount
    cols[13] = pay14
    cols[15] = amount14
    return "\t".join(cols)


# 헤더 2줄(채널그룹 등 — col0 가 채널값이 아니라 자동 skip).
_HEADER = "\t".join(["채널그룹", "채널", "nt_medium", "nt_detail", "nt_keyword",
                     "유입(전체)", "유입수", "c7", "c8", "결제수", "c10", "결제금액",
                     "c12", "결제수14", "c14", "결제금액14"])
_HEADER2 = "\t".join(["속성"] + ["-"] * 15)

# 시경 표 구조 재현(전체 그랜드토탈 + 모바일/PC 기기행, 매체 대소문자 혼용·빈 keyword·형식 혼용).
# +14일 기여도추정: 전체 결제금액14 = 3,367,195(마지막클릭 2,919,930 대비 +15%).
_SAMPLE = "\n".join([
    _HEADER, _HEADER2,
    _row("전체", "", "", "", "1,585", "113", "2,919,930", "130", "3,367,195"),
    _row("모바일", "blog", "P01filter2ea", "에어컨필터", "150", "5", "100,000", "6", "120,000"),
    _row("PC", "blog", "P01filter2ea", "에어컨필터", "79", "4", "80,000", "5", "95,000"),
    _row("모바일", "blog", "260622_와이퍼_렉스턴스포츠", "와이퍼", "76", "19", "300,000", "22", "350,000"),
    _row("모바일", "REVU", "P02wiper", "", "200", "10", "250,000", "12", "300,000"),
    _row("PC", "revu", "P02wiper", "-", "86", "4", "90,000", "5", "110,000"),
    "깨진\t줄",                                                       # 컬럼 부족 → skip
])


def test_summary_from_total_row():
    out = parse_smartstore_table(_SAMPLE)
    s = out["summary"]
    assert s["inflow"] == 1585
    assert s["pay"] == 113
    assert round(s["pay_rate"], 1) == 7.1
    assert s["amount"] == 2919930


def test_by_medium_case_insensitive_merge():
    out = parse_smartstore_table(_SAMPLE)
    by = {m["medium"]: m for m in out["by_medium"]}
    assert by["blog"]["inflow"] == 305 and by["blog"]["pay"] == 28
    assert by["revu"]["inflow"] == 286 and by["revu"]["pay"] == 14   # REVU+revu 통합
    # rate 내림차순 정렬
    rates = [m["rate"] for m in out["by_medium"]]
    assert rates == sorted(rates, reverse=True)


def test_by_campaign_merges_mobile_and_pc():
    out = parse_smartstore_table(_SAMPLE)
    camp = {c["campaign"]: c for c in out["by_campaign"]}
    assert camp["P01filter2ea"]["inflow"] == 229   # 150(모바일)+79(PC)
    assert camp["P01filter2ea"]["pay"] == 9
    rates = [c["rate"] for c in out["by_campaign"]]
    assert rates == sorted(rates, reverse=True)


def test_warnings_three_kinds_detected():
    out = parse_smartstore_table(_SAMPLE)
    w = " | ".join(out["warnings"])
    assert "대소문자" in w        # REVU/revu
    assert "비어있음" in w        # keyword "" / "-"
    assert "형식" in w            # 언더스코어형 + 비언더스코어형 혼용
    assert len(out["warnings"]) >= 3


def test_by_product_new_format_token():
    """신규형 detail 만 있는 표 → 제품=토큰[1] 로 by_product 집계."""
    raw = "\n".join([
        _HEADER,
        _row("전체", "", "", "", "76", "19", "300,000"),
        _row("모바일", "blog", "260622_와이퍼_렉스턴스포츠", "와이퍼", "76", "19", "300,000"),
    ])
    out = parse_smartstore_table(raw)
    prod = {p["product"]: p for p in out["by_product"]}
    assert "와이퍼" in prod
    assert prod["와이퍼"]["inflow"] == 76 and prod["와이퍼"]["pay"] == 19


def test_by_product_empty_when_no_new_format():
    """신규형 detail 이 하나도 없으면 by_product 는 빈 리스트."""
    raw = "\n".join([
        _HEADER,
        _row("모바일", "blog", "P01filter2ea", "에어컨필터", "150", "5", "100,000"),
    ])
    out = parse_smartstore_table(raw)
    assert out["by_product"] == []


def test_summary_falls_back_to_sum_when_no_total_row():
    """'전체' 행이 없으면 기기(모바일/PC)행 합산으로 summary."""
    raw = "\n".join([
        _HEADER,
        _row("모바일", "blog", "P01filter2ea", "에어컨필터", "150", "5", "100,000"),
        _row("PC", "blog", "P01filter2ea", "에어컨필터", "79", "4", "80,000"),
    ])
    out = parse_smartstore_table(raw)
    assert out["summary"]["inflow"] == 229 and out["summary"]["pay"] == 9


def test_num_parsing_helpers():
    assert _num("1,585") == 1585.0
    assert _num("7.13%") == 7.13
    assert _num("") == 0.0
    assert _num("abc") == 0.0
    assert _is_num("1,585") is True
    assert _is_num("") is False
    assert _is_num("-") is False


def test_broken_and_short_lines_skipped():
    raw = "\n".join([
        _HEADER,
        "짧은\t줄",                                   # 컬럼 부족
        "그냥아무텍스트탭없음",                         # 탭 없음
        _row("기타", "blog", "X", "kw", "10", "1", "1,000"),  # col0 채널 아님 → skip
        _row("모바일", "blog", "P01filter2ea", "에어컨필터", "150", "5", "100,000"),
    ])
    out = parse_smartstore_table(raw)
    # 유효 기기행 1개만 집계됨
    assert out["summary"]["inflow"] == 150
    assert len(out["by_campaign"]) == 1


def test_rate_zero_division_guard():
    raw = "\n".join([
        _HEADER,
        _row("전체", "", "", "", "0", "0", "0"),
    ])
    out = parse_smartstore_table(raw)
    assert out["summary"]["pay_rate"] == 0.0


# ── rows 키(엑셀 시트1 재현용) ──────────────────────────────────────────────
def test_rows_includes_total_and_device_with_source():
    # source(col1) 가 채워진 표 — _row 헬퍼는 col1 을 "-" 로 두므로 직접 구성.
    def _r(channel, source, medium, detail, kw, inflow, pay, amount):
        cols = ["-"] * 12
        cols[0], cols[1], cols[2], cols[3], cols[4] = channel, source, medium, detail, kw
        cols[6], cols[9], cols[11] = inflow, pay, amount
        return "\t".join(cols)

    raw = "\n".join([
        _HEADER,
        _r("전체", "naver.blog", "", "", "", "1,585", "113", "2,919,930"),
        _r("모바일", "naver.blog", "blog", "P01filter2ea", "에어컨필터", "150", "5", "100,000"),
    ])
    out = parse_smartstore_table(raw)
    rows = out["rows"]
    assert len(rows) == 2                       # 전체 + 기기행 모두 포함
    assert rows[0]["channel"] == "전체" and rows[0]["inflow"] == 1585
    assert rows[0]["source"] == "naver.blog"    # col1 채워짐
    assert round(rows[1]["pay_rate"], 4) == round(5 / 150 * 100, 4)


# ── +14일 기여도추정 보조 지표 ──────────────────────────────────────────────
def test_summary_includes_pay14_amount14():
    out = parse_smartstore_table(_SAMPLE)
    s = out["summary"]
    assert s["pay14"] == 130
    assert s["amount14"] == 3367195
    # 메인(마지막클릭)은 그대로 — 회귀 차단
    assert s["pay"] == 113 and s["amount"] == 2919930


def test_by_medium_and_campaign_pay14_aggregated():
    out = parse_smartstore_table(_SAMPLE)
    by = {m["medium"]: m for m in out["by_medium"]}
    assert by["blog"]["pay14"] == 33 and by["blog"]["amount14"] == 565000
    assert by["revu"]["pay14"] == 17 and by["revu"]["amount14"] == 410000
    camp = {c["campaign"]: c for c in out["by_campaign"]}
    assert camp["P01filter2ea"]["pay14"] == 11
    assert camp["P01filter2ea"]["amount14"] == 215000


def test_rows_include_pay14_amount14():
    out = parse_smartstore_table(_SAMPLE)
    grand = out["rows"][0]
    assert grand["channel"] == "전체"
    assert grand["pay14"] == 130 and grand["amount14"] == 3367195


# ── 엑셀 내보내기 ────────────────────────────────────────────────────────────
def _summary_cells(ws):
    """분석결과 요약 항목명 → 값 셀(object) 매핑."""
    out = {}
    for row in ws.iter_rows():
        if row and row[0].value in (
                "총 유입", "총 결제", "결제율_표시", "결제율_정확", "결제금액",
                "결제수(+14일)", "결제금액(+14일)"):
            out[row[0].value] = row[1]
    return out


def test_build_analytics_xlsx_two_sheets_and_numbers():
    result = parse_smartstore_table(_SAMPLE)
    data = build_analytics_xlsx(result)
    assert data and len(data) > 0

    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["원본붙여넣기", "분석결과"]

    # 시트1: 전체행 유입 1585 가 어딘가 존재(정수 유지)
    ws1 = wb["원본붙여넣기"]
    vals1 = [c.value for row in ws1.iter_rows() for c in row]
    assert 1585 in vals1

    # 시트2 요약: 정수는 그대로, 결제율은 ★비율(<1.0) + % 서식
    ws2 = wb["분석결과"]
    sm = _summary_cells(ws2)
    assert sm["총 유입"].value == 1585          # 비율변환 안 됨
    assert sm["총 결제"].value == 113
    assert sm["결제금액"].value == 2919930
    show, exact = sm["결제율_표시"], sm["결제율_정확"]
    assert show.value < 1.0 and "%" in show.number_format
    assert exact.value < 1.0 and "%" in exact.number_format
    assert round(show.value, 3) == 0.071        # 7.1% 비율
    assert round(exact.value, 6) == 0.071293    # 113/1585 풀소수 비율
    # +14일 보조: 요약에 결제금액(+14일) 행 존재(정수)
    assert sm["결제금액(+14일)"].value == 3367195


def test_xlsx_has_14day_columns():
    """시트1 헤더에 결제금액_14일 컬럼이 추가되고 값이 정수로 기록된다."""
    result = parse_smartstore_table(_SAMPLE)
    wb = load_workbook(BytesIO(build_analytics_xlsx(result)))
    ws1 = wb["원본붙여넣기"]
    header = [c.value for c in next(ws1.iter_rows())]
    assert "결제수_14일" in header and "결제금액_14일" in header
    # 전체행 결제금액_14일 값이 존재
    vals1 = [c.value for row in ws1.iter_rows() for c in row]
    assert 3367195 in vals1


def test_build_analytics_xlsx_header_styled_bold():
    """표 헤더셀(매체 집계의 '유입')이 굵게 강조된다(서식 회귀 차단)."""
    result = parse_smartstore_table(_SAMPLE)
    wb = load_workbook(BytesIO(build_analytics_xlsx(result)))
    ws2 = wb["분석결과"]
    found_bold = False
    for row in ws2.iter_rows():
        if row and row[0].value == "매체":          # 매체 집계 헤더행
            for cell in row:
                if cell.value == "유입":
                    found_bold = cell.font.bold is True
    assert found_bold, "매체 표 헤더 '유입' 셀이 bold 아님"
